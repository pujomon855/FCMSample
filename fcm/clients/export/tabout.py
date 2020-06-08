# -*- coding: utf-8 -*-

import openpyxl
import os
from enum import IntEnum, auto

from fcm import settings
from ..models import Client, ClientTrade, TradeType
from ..models import CostType
from ..utils.clients import trade_types_sdb_to_str, trade_types_to_str


CLIENT_TABLE_FILE_PATH = os.path.join(settings.BASE_DIR, 'clients/static/clients/excel/FIXClientTable.xlsx')
CLIENT_TABLE_FILE_NAME = 'FIXClientTable.xlsx'


START_ROW = 3
COST_START_COL = 11

NUM_FIX_LINE = 6  # FIX&Line列がいくつあるか

COST_TYPE_FIX_AND_LINE = CostType.objects.get(name='FIX and Line')
COST_TYPE_LINE = CostType.objects.get(name='Line')
COST_TYPE_OMS = CostType.objects.get(name='OMS')
COST_TYPE_EMS = CostType.objects.get(name='EMS')
COST_TYPE_IOI = CostType.objects.get(name='IOI')


class BaseColumn(IntEnum):
    def _generate_next_value_(self, start, count, last_values):
        """
        auto()から呼ばれた際、最後に提供した int の次から順々に int を提供する。
        この動作は実装詳細であり変更される可能性があるために明示的に実装。
        """
        for last_value in reversed(last_values):
            try:
                return last_value + 1
            except TypeError:
                pass
        else:
            return start


class FIXColumn(BaseColumn):
    NO = 1
    CONN_START_DATE = auto()
    OMS = auto()
    CLIENT_NAME = auto()
    VIEW_CODE = auto()
    SESSION_NAME = auto()
    FIX_CODE = auto()
    TRADE_TYPE_SDB = auto()
    TRADE_TYPE = auto()


class CostColumn(BaseColumn):
    VENDOR_NAME = 0
    PRODUCT = auto()
    CHANGE_TYPE = auto()
    CHANGE = auto()
    CURRENCY = auto()


def create_client_table_xl():
    """
    顧客一覧エクセルを作成する。

    :return: (顧客一覧wb(未保存), ファイル名)
    """
    wb = openpyxl.load_workbook(CLIENT_TABLE_FILE_PATH)
    ws = wb.active

    clients = Client.objects.order_by('name')
    row_idx = START_ROW
    for client in clients:
        classifiers = client.clientclassifier_set.order_by('view')
        for classifier in classifiers:
            code_sessions = classifier.codesession_set.all()
            trade_types_sdb = ClientTrade.objects.using('limit').get(client_id=classifier.identifier)
            trade_types_sdb_str = trade_types_sdb_to_str(trade_types_sdb)
            for code_session in code_sessions:
                trade_types = TradeType.objects.filter(code=classifier, session=code_session.session)
                _put_fix_data(ws, row_idx, client, classifier, code_session, trade_types_sdb_str, trade_types)
                _put_cost_data(ws, row_idx, code_session)
                row_idx += 1

    return wb, CLIENT_TABLE_FILE_NAME


def _put_fix_data(ws, row_idx, client, classifier, code_session, trade_types_sdb_str, trade_types):
    ws.cell(row=row_idx, column=FIXColumn.NO, value=(row_idx - START_ROW + 1))
    ws.cell(row=row_idx, column=FIXColumn.CONN_START_DATE, value=code_session.connection_start_date)
    ws.cell(row=row_idx, column=FIXColumn.OMS, value='')
    ws.cell(row=row_idx, column=FIXColumn.CLIENT_NAME, value=client.name)
    ws.cell(row=row_idx, column=FIXColumn.VIEW_CODE, value=classifier.view)
    ws.cell(row=row_idx, column=FIXColumn.SESSION_NAME, value=code_session.session.name)
    ws.cell(row=row_idx, column=FIXColumn.FIX_CODE, value=classifier.code)
    ws.cell(row=row_idx, column=FIXColumn.TRADE_TYPE_SDB, value=trade_types_sdb_str)
    ws.cell(row=row_idx, column=FIXColumn.TRADE_TYPE, value=trade_types_to_str(trade_types))


def _put_cost_data(ws, row_idx, code_session):
    cost_data_set = code_session.cost_set.all()
    num_cost_cols = len(CostColumn)
    _put_fix_lines(ws, row_idx, cost_data_set.filter(cost_type=COST_TYPE_FIX_AND_LINE), num_cost_cols)

    start_col_idx = COST_START_COL + NUM_FIX_LINE * num_cost_cols
    _put_cost_data_ind_list(ws, row_idx, start_col_idx, cost_data_set.filter(cost_type=COST_TYPE_LINE))

    start_col_idx += num_cost_cols
    _put_cost_data_ind_list(ws, row_idx, start_col_idx, cost_data_set.filter(cost_type=COST_TYPE_OMS))

    start_col_idx += num_cost_cols
    _put_cost_data_ind_list(ws, row_idx, start_col_idx, cost_data_set.filter(cost_type=COST_TYPE_EMS))

    start_col_idx += num_cost_cols
    _put_cost_data_ind_list(ws, row_idx, start_col_idx, cost_data_set.filter(cost_type=COST_TYPE_IOI))


def _put_fix_lines(ws, row_idx, fix_lines, num_cost_cols):
    for n, fix_line in enumerate(fix_lines[:NUM_FIX_LINE]):
        if n >= NUM_FIX_LINE:
            break
        start_col_idx = COST_START_COL + n * num_cost_cols
        _put_cost_data_ind(ws, row_idx, start_col_idx, fix_line)


def _put_cost_data_ind_list(ws, row_idx, start_col_idx, data_list):
    """
    data_listは大きさが0か1のiterable。1以上の時は先頭のデータを出力する
    """
    if data_list:
        _put_cost_data_ind(ws, row_idx, start_col_idx, data_list[0])


def _put_cost_data_ind(ws, row_idx, start_col_idx, data):
    ws.cell(row=row_idx, column=(CostColumn.VENDOR_NAME + start_col_idx), value=data.vendor.name)
    product_str = ''
    if data.product and data.handlinst:
        product_str = f'{data.product}({data.handlinst})'
    elif data.product:
        product_str = data.product.name
    elif data.handlinst:
        product_str = data.handlinst.name
    ws.cell(row=row_idx, column=(CostColumn.PRODUCT + start_col_idx), value=product_str)
    ws.cell(row=row_idx, column=(CostColumn.CHANGE_TYPE + start_col_idx), value=data.change_type.name)
    ws.cell(row=row_idx, column=(CostColumn.CHANGE + start_col_idx), value=data.change)
    ws.cell(row=row_idx, column=(CostColumn.CURRENCY + start_col_idx),
            value=_to_str_null_safe(data.currency))


def _to_str_null_safe(val):
    if val:
        return val.name
    else:
        return ''
